import sys
import os
import time
import numpy as np
import openpyxl

from PyQt5.QtWidgets import QApplication

from Environment import TrafficDRL_Env
from Render_Widget import mainWidget
from stable_baselines3 import PPO
from stable_baselines3.ppo import MlpPolicy
from stable_baselines3.common.callbacks import CheckpointCallback

from Environment import  all_stay_data, all_wait_data, red_light_time_list
#from Environment import flow_1, flow_2, flow_3, flow_4

class TrafficDRL():
    def __init__(self):
        self.widget = mainWidget()
        self.n_steps = 1200
        self.n_train_episodes = 500 # change to 3000 if es=2
        self.n_episode_per_callback = 50 # change to 150 if es=2
        self.save_path = './models4/rf5,es2/ent_coef=0.0/gamma=0.9999/'
        self.stay_excel_name = "stay fixed_time.xlsx"
        self.wait_excel_name = "wait fixed time.xlsx"

        self.use_fixed_time_system = False
        

        rf = 5 # Option of 0 ~ 7, each represents different reward function.
               # (0): The negative of the average time of cars staying in the environment, 
               # minus the traffic signal penalty.
               # (1): The negative of the total time of cars staying in the environment, 
               # minus the traffic signal penalty.
               # (2): The negative of the average time of waiting in front of traffic signals, 
               # minus the traffic signal penalty.
               # (3): The negative of the total time of waiting in front of traffic signals, 
               # minus the traffic signal penalty.
               # (4~7): Delta time equivalent of (0~3)

        env_sys = 2 # Option of 1 ~ 3, represents different environment scale.
                # (1): 1 crossroads, 1 master signals, 2 paths. each has a length of 400 meters.
                # (2): 4 crossroads, 4 master signals, 4 paths. each has a length of 600 meters.
                # (3): 9 crossroads, 9 master signals, 6 paths. each has a length of 600 meters.

        self.env = TrafficDRL_Env(
            reward_function=rf,
            env_sys=env_sys,
            render_scene=self.widget.viewTab.scene,
            n_steps=self.n_steps
        )

        self.checkpoint_callback = CheckpointCallback(save_freq=self.n_steps*self.n_episode_per_callback, save_path=self.save_path,
                                         name_prefix='a_')
        
        if not self.use_fixed_time_system:
            self.model = PPO(
                MlpPolicy,
                self.env,
                verbose=1,
                n_steps=self.n_steps,
                learning_rate=3e-4,
                ent_coef=0,
                gamma=0.95,
                tensorboard_log="./tensorboard/"
            )
        else:
            self.model = fixed_time_system_model()
            

    def test(self, flow=None, render=False, fixed_time_red_light_time=None):
        delay = 0.001
        i = 0
        if render:
            self.env.render()
        obs = self.env.reset(isTest=True)
        self.test_done = False
        while not self.test_done:
            action, _states = self.model.predict(obs, deterministic=True)
            print(action)
            obs, reward, self.test_done, info = self.env.step(action, delay)
            i+=1
        print(i)
        self.env.render(close=True)

    def Training(self):
        drl_app.model.save(drl_app.save_path+'/a__0_steps')
        drl_app.model.learn(drl_app.n_train_episodes*drl_app.n_steps, drl_app.checkpoint_callback)
    
    def Testing(self):
        drl_app.model.set_env(drl_app.env)
        model = '/a__600000_steps'
        drl_app.model = PPO.load(drl_app.save_path + str(model))
        drl_app.test(flow=[5, 10, 20, 10], fixed_time_red_light_time=12)

    def collecting_data_for_stay_and_wait(self):
        #excel for stay
        
        model_num = 11
        tests_per_step = 60000
        test_time = 5
        avg_list = [0] * (model_num+1)
        avg_list[0] = "fs1"
        lists = [0] * (model_num)
        all_avg_data = []
        tlist = []
        tblanklist = []
        tblanklist.append(" ")
        percentage = 1
        wb1 = openpyxl.Workbook()
        ws1 = wb1.active

        for i in range(model_num):
            tlist.append("t" +str(i))
            tblanklist.append("t" +str(i))
        ws1.append(tlist)
        
        for i, flows in enumerate(test_flow_sets):
            for j in range(test_time):
                for k in range(model_num):
                    model = '/a__' + str(k*tests_per_step) + '_steps.zip'
                    drl_app.model = PPO.load(drl_app.save_path + str(model))
                    drl_app.test(flow=flows)
                    print("------"+str(round(percentage/((model_num*test_time)*len(test_flow_sets))*100 , 4))+" %------")
                    percentage += 1
            
        number = 2
        for i, data in enumerate(all_stay_data):
            j = ((i+model_num)%model_num)+1
            lists[j-1] = data
            avg_list[j] += data

            if (i+1)%model_num == 0:
                ws1.append(lists)

            if (i+1)%(model_num*test_time) == 0:
                for i in range(1, model_num+1):
                    avg_list[i]/=test_time
                ws1.append(avg_list)
                all_avg_data.append(avg_list)
                avg_list = [0] * (model_num+1)
                avg_list[0] = "fs"+str(number)
                ws1.append(["-"])
                number += 1

        ws1.append(tblanklist)
        for data in all_avg_data:
            ws1.append(data)

        c1 = openpyxl.chart.LineChart()
        data = openpyxl.chart.Reference(ws1, min_col=1, min_row=2+len(test_flow_sets)*(test_time+2), max_col=model_num+1, max_row=2+len(test_flow_sets)+len(test_flow_sets)*(test_time+2)) 
        c1.add_data(data, titles_from_data=True)

        ws1.add_chart(c1, "O1")
        wb1.save(self.stay_excel_name)
        
        #excel for wait

        avg_list = [0] * (model_num+1)
        avg_list[0] = "fs1"
        lists = [0] * (model_num)
        all_avg_data = []
        tlist = []
        tblanklist = []
        tblanklist.append(" ")
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active

        for i in range(model_num):
            tlist.append("t" +str(i))
            tblanklist.append("t" +str(i))
        ws2.append(tlist)

        number = 2
        for i, data in enumerate(all_wait_data):
            j = ((i+model_num)%model_num)+1
            lists[j-1] = data
            avg_list[j] += data

            if (i+1)%model_num == 0:
                ws2.append(lists)

            if (i+1)%(model_num*test_time) == 0:
                for i in range(1, model_num+1):
                    avg_list[i]/=test_time
                ws2.append(avg_list)
                all_avg_data.append(avg_list)
                avg_list = [0] * (model_num+1)
                avg_list[0] = "fs"+str(number)
                ws2.append(["-"])
                number += 1

        ws2.append(tblanklist)
        for data in all_avg_data:
            ws2.append(data)

        c1 = openpyxl.chart.LineChart()
        data = openpyxl.chart.Reference(ws2, min_col=1, min_row=2+len(test_flow_sets)*(test_time+2), max_col=model_num+1, max_row=6+len(test_flow_sets)*(test_time+2)) 
        c1.add_data(data, titles_from_data=True)

        ws2.add_chart(c1, "O1")
        wb2.save(self.wait_excel_name)

        
    def collecting_red_light_time(self):
        #excel for red light time
        red_light_tot_time_list = []
        list_ = [0] * 1
        wb3 = openpyxl.Workbook()
        ws3 = wb3.active

        for i, flows in enumerate(test_flow_sets):
            list_ = ["flow" + str(i+1)]
            self.env.all_light_time_data.append(list_)
            model = '/a__' + str(600000) + '_steps.zip'
            drl_app.model = PPO.load(drl_app.save_path + str(model))
            drl_app.test(flow=flows)
        
        print(self.env.all_light_time_data)

        for list_data in self.env.all_light_time_data:
            ws3.append(list_data)
        
        wb3.save(  "red light tot time test.xlsx")

    def real_time(self, render=False):
        if render:
            self.env.render()

        drl_app.model.set_env(drl_app.env)
        model = '/a__600000_steps'
        drl_app.model = PPO.load(drl_app.save_path + str(model))

        obs = self.env.reset([0,0,0,0], isTest=True, run_forever=True, delay=True)
        self.test_done = False
        while 1:
            fixed_flow = drl_app.widget.slider.get_realtime_flow()
            delay = drl_app.widget.slider.get_step_sleep()
            drl_app.env.change_flow(fixed_flow)
            action, _states = self.model.predict(obs, deterministic=True)
            obs, reward, self.test_done, info = self.env.step(action, delay)

    def collect_action_and_avg_staytime(self, render=False):
        wb5 = openpyxl.Workbook()
        ws5 = wb5.active

        drl_app.model.set_env(drl_app.env)
        model = '/a__600000_steps'
        drl_app.model = PPO.load(drl_app.save_path + str(model))


        number_list = []
        action_list = []

        for i in range(len(self.env.master_signals)):
            action_list.append([])
        last_avg_stay_list = []
        number = 0

        for i in range(400):
            number_list.append(i+1)

        delay = 0.0
        i = 0
        if render:
            self.env.render()
        obs = self.env.reset(fixed_flow=[10, 10, 10, 10], isTest=True)
        self.test_done = False
        for i in range(400):
            if i == 200:
                self.env.change_flow([15, 5 ,15, 5])
            action, _states = self.model.predict(obs, deterministic=True)
            obs, reward, self.test_done, info = self.env.step(action, delay)
            print(action)
            for j in range(len(action.tolist())):
                action_list[j].append(action.tolist()[j])
            last_avg_stay_list.append(self.env.get_cur_avg_stay()) 
            i+=1
        print(i)
        self.env.render(close=True)

        ws5.append(number_list)
        for data in action_list:
            ws5.append(data)
        ws5.append(last_avg_stay_list)
        wb5.save( "action_and_avg_staytime [10, 10, 10, 10]  [15, 5, 15, 5] sys 2 trained gamma=0.95.xlsx")

    


test_flow_sets = [[10, 10],
                  [5, 15],
                  [15, 5],
                  [20, 20]]

test_flow_sets = [[10, 10, 10, 10],
                  [5, 5, 15, 15],
                  [15, 15, 5, 5],
                  [15, 5, 15, 5],
                  [5, 15, 5, 15],
                  [15, 5, 15, 5],
                  [5, 15, 5, 15],
                  [20, 20, 20, 20]]

test_flow_sets = [[15, 15, 5, 5],
                  [15, 5, 15, 5]]

if __name__ == '__main__':
    app = QApplication(sys.argv)
    drl_app = TrafficDRL()
    drl_app.widget.show()

    #=================================

    # real time test
    #drl_app.real_time(render=True)
    #=================================

    ''' train and test '''
    # For Training
    #drl_app.Training()

    # For Testing

    '''
    if drl_app.use_fixed_time_system:
        drl_app.test(fixed_time_red_light_time=18)
    else:
        drl_app.Testing()
    '''

    #=================================

    ''' collecting datas '''
    # collecting data for stay and wait
    #drl_app.collecting_data_for_stay_and_wait()

    # collecting red light time data
    #drl_app.collecting_red_light_time()

    #=================================

    #drl_app.fixed_time_system_model_collect_data()

    drl_app.collect_action_and_avg_staytime(render=False)
    print("finish")
    
    os._exit(app.exec_())